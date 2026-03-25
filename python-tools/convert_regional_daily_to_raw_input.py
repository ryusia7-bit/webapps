from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent

RAW_GROUP_HEADERS = [
    "공통", "공통", "공통", "공통", "공통", "공통", "공통",
    "병원", "병원", "병원", "병원",
    "시설", "시설", "시설", "시설",
    "서비스", "서비스", "서비스", "서비스", "서비스", "서비스", "서비스",
    "공적서비스", "공적서비스", "공적서비스",
]

RAW_DETAIL_HEADERS = [
    "번호", "날짜", "담당자", "이름", "생년월일", "성별", "상담유형",
    "입원유형", "병원명", "입원여부", "퇴원후연계",
    "입소유형", "시설명", "입퇴소여부", "퇴소후조치",
    "병원명", "주거지원", "물품제공", "방문상담", "외래진료", "투약관리", "기타",
    "지원항목", "신청기관", "지원시기",
]

RAW_COLUMN_KEYS = [
    "공통_번호", "공통_날짜", "공통_담당자", "공통_이름", "공통_생년월일", "공통_성별", "공통_상담유형",
    "병원_입원유형", "병원_병원명", "병원_입원여부", "병원_퇴원후연계",
    "시설_입소유형", "시설_시설명", "시설_입퇴소여부", "시설_퇴소후조치",
    "서비스_병원명", "서비스_주거지원", "서비스_물품제공", "서비스_방문상담", "서비스_외래진료", "서비스_투약관리", "서비스_기타",
    "공적서비스_지원항목", "공적서비스_신청기관", "공적서비스_지원시기",
]

RAW_COLUMN_WIDTHS = [
    12, 12, 12, 12, 12, 8, 12,
    12, 18, 12, 20,
    12, 20, 12, 20,
    18, 10, 10, 10, 10, 10, 28,
    18, 18, 12,
]

SECTION_FIELDS = {
    "hospital": ["병원_입원유형", "병원_병원명", "병원_입원여부", "병원_퇴원후연계"],
    "facility": ["시설_입소유형", "시설_시설명", "시설_입퇴소여부", "시설_퇴소후조치"],
    "service": [
        "서비스_병원명", "서비스_주거지원", "서비스_물품제공",
        "서비스_방문상담", "서비스_외래진료", "서비스_투약관리", "서비스_기타",
        "공적서비스_지원항목", "공적서비스_신청기관", "공적서비스_지원시기",
    ],
}

GROUP_RANGES = [
    ("공통", 1, 7, "DDEBF7"),
    ("병원", 8, 11, "FCE4D6"),
    ("시설", 12, 15, "E2F0D9"),
    ("서비스", 16, 22, "FFF2CC"),
    ("공적서비스", 23, 25, "EADCF8"),
]

PUBLIC_SERVICE_KEYWORDS = {
    "말소복원": ["말소복원", "말소 복원", "주민등록복원"],
    "주민등록": ["주민등록", "신분증", "재발급", "십지문", "신원조회"],
    "수급상담": ["수급상담", "수급 안내", "수급", "기초생활"],
    "긴급지원": ["긴급지원"],
}

HOSPITAL_NAME_PATTERN = re.compile(r"([가-힣A-Za-z0-9\s·]+?(?:병원|정신건강의학과))")


@dataclass
class LogEntry:
    level: str
    source_sheet: str
    source_row: int
    message: str


@dataclass
class ImportRecord:
    data: dict[str, str] = field(default_factory=dict)
    source_refs: list[str] = field(default_factory=list)


class RecordStore:
    def __init__(self) -> None:
        self.records: list[ImportRecord] = []
        self.by_exact_key: defaultdict[tuple[str, str, str, str], list[int]] = defaultdict(list)
        self.by_person_key: defaultdict[tuple[str, str, str], list[int]] = defaultdict(list)

    def _empty_record(self) -> ImportRecord:
        return ImportRecord(data={key: "" for key in RAW_COLUMN_KEYS})

    def create(self, exact_key: tuple[str, str, str, str], source_ref: str) -> ImportRecord:
        record = self._empty_record()
        index = len(self.records)
        self.records.append(record)
        self.by_exact_key[exact_key].append(index)
        self.by_person_key[exact_key[:3]].append(index)
        record.source_refs.append(source_ref)
        return record

    def find_attachable(
        self,
        exact_key: tuple[str, str, str, str],
        section_name: str,
    ) -> ImportRecord | None:
        fields = SECTION_FIELDS[section_name]
        for lookup_key in (exact_key, None):
            indexes = self.by_exact_key.get(exact_key, []) if lookup_key is exact_key else self.by_person_key.get(exact_key[:3], [])
            for index in reversed(indexes):
                record = self.records[index]
                if all(not normalize_text(record.data.get(field)) for field in fields):
                    return record
        return None

    def finalize_numbers(self) -> None:
        for index, record in enumerate(self.records, start=1):
            record.data["공통_번호"] = f"LEGACY-{index:05d}"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_date_text(value: object) -> str:
    if value in (None, ""):
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    text = normalize_text(value)
    if not text:
        return ""

    for pattern in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, pattern).strftime("%Y-%m-%d")
        except ValueError:
            continue

    match = re.match(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if match:
        year, month, day = map(int, match.groups())
        return f"{year:04d}-{month:02d}-{day:02d}"

    return ""


def normalize_birth_text(value: object) -> str:
    if value in (None, ""):
        return ""

    if isinstance(value, datetime):
        return value.strftime("%y%m%d")
    if isinstance(value, date):
        return value.strftime("%y%m%d")

    text = normalize_text(value)
    digits = re.sub(r"\D", "", text)
    if len(digits) >= 8 and digits.startswith(("19", "20")):
        candidate = digits[2:8]
        return candidate if is_valid_birth_yymmdd(candidate) else ""
    if len(digits) >= 6:
        candidate = digits[:6]
        return candidate if is_valid_birth_yymmdd(candidate) else ""
    return ""


def is_valid_birth_yymmdd(value: str) -> bool:
    digits = re.sub(r"\D", "", normalize_text(value))
    if len(digits) != 6:
        return False

    year = 2000 + int(digits[0:2])
    month = int(digits[2:4])
    day = int(digits[4:6])

    try:
        parsed = date(year, month, day)
    except ValueError:
        return False

    return (
        parsed.year == year and
        parsed.month == month and
        parsed.day == day
    )


def normalize_gender(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if text.startswith("남"):
        return "남"
    if text.startswith("여"):
        return "여"
    return text[:1]


def normalize_counsel_type(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    lowered = text.lower()
    if "알코올" in text or "알콜" in text or "alcohol" in lowered:
        return "알코올"
    if "정신" in text:
        return "정신"
    return "기타"


def normalize_hospital_type(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if "자의" in text:
        return "자의입원"
    if "동의" in text:
        return "동의입원"
    if "응급" in text:
        return "응급입원"
    if "행정" in text:
        return "행정입원"
    if "보호진단" in text:
        return "행정입원"
    if "보호입원" in text:
        return "동의입원"
    return "기타"


def normalize_hospital_status(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if "퇴원" in text:
        return "퇴원"
    if "입원" in text:
        return "입원"
    return "기타"


def normalize_facility_type(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if "자활" in text:
        return "자활"
    if "재활" in text:
        return "재활"
    if "요양" in text:
        return "요양"
    if "일시" in text or "보호" in text:
        return "일시보호"
    return "기타"


def normalize_facility_status(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if "퇴소" in text:
        return "퇴소"
    if "입소" in text:
        return "입소"
    return "기타"


def mark_checkbox(record: ImportRecord, field_key: str) -> None:
    record.data[field_key] = "TRUE"


def append_text(record: ImportRecord, field_key: str, text: str) -> None:
    text = normalize_text(text)
    if not text:
        return
    current = normalize_text(record.data.get(field_key))
    if not current:
        record.data[field_key] = text
        return
    existing_parts = {part.strip() for part in current.split(" / ") if part.strip()}
    if text not in existing_parts:
        record.data[field_key] = f"{current} / {text}"


def set_if_blank(record: ImportRecord, field_key: str, value: str) -> None:
    value = normalize_text(value)
    if value and not normalize_text(record.data.get(field_key)):
        record.data[field_key] = value


def build_keys(date_text: str, worker: str, name: str, birth: str) -> tuple[tuple[str, str, str, str], tuple[str, str, str]]:
    exact_key = (date_text, normalize_text(name), normalize_text(birth), normalize_text(worker))
    person_key = exact_key[:3]
    return exact_key, person_key


def extract_hospital_name(*texts: object) -> str:
    for text in texts:
        cleaned = normalize_text(text)
        if not cleaned:
            continue
        match = HOSPITAL_NAME_PATTERN.search(cleaned)
        if match:
            candidate = match.group(1).strip()
            if any(fragment in candidate for fragment in ["및 병원", "병원진료", "병원 연계", "병원 외래"]):
                continue
            return candidate
    return ""


def infer_public_service_items(*texts: object) -> list[str]:
    source_text = " ".join(normalize_text(text) for text in texts if normalize_text(text))
    items: list[str] = []
    for label, keywords in PUBLIC_SERVICE_KEYWORDS.items():
        if any(keyword in source_text for keyword in keywords):
            items.append(label)
    if "공적서비스" in source_text and "공적서비스" not in items:
        items.append("공적서비스")
    return items


def infer_public_service_org(referral_org: str, detail: str) -> str:
    source = " ".join(filter(None, [normalize_text(referral_org), normalize_text(detail)]))
    if "주민센터" in source:
        return "주민센터"
    if "구청" in source:
        return "구청"
    return ""


def summarize_unstructured_service(result: str, request: str, detail: str) -> str:
    result = normalize_text(result)
    request = normalize_text(request)
    detail = normalize_text(detail)

    if result in {"주거지원", "외래진료"}:
        return ""

    parts = []
    if result:
        parts.append(result)
    if request and request not in parts:
        parts.append(request)
    if detail:
        parts.append(detail)
    return " | ".join(parts)


def apply_common_fields(
    record: ImportRecord,
    *,
    date_text: str,
    worker: str,
    name: str,
    birth: str,
    gender: str,
    counsel_type: str,
) -> None:
    set_if_blank(record, "공통_날짜", date_text)
    set_if_blank(record, "공통_담당자", worker)
    set_if_blank(record, "공통_이름", name)
    set_if_blank(record, "공통_생년월일", birth)
    set_if_blank(record, "공통_성별", gender)
    set_if_blank(record, "공통_상담유형", counsel_type)


def get_or_create_record(
    store: RecordStore,
    *,
    section_name: str,
    source_ref: str,
    exact_key: tuple[str, str, str, str],
) -> ImportRecord:
    record = store.find_attachable(exact_key, section_name)
    if record is None:
        record = store.create(exact_key, source_ref)
    else:
        record.source_refs.append(source_ref)
    return record


def load_source_workbook(input_path: Path):
    return load_workbook(input_path, data_only=True)


def ingest_counsel_sheet(workbook, sheet_name: str, store: RecordStore, logs: list[LogEntry], counters: Counter) -> None:
    ws = workbook[sheet_name]
    for row_idx in range(2, ws.max_row + 1):
        name = normalize_text(ws.cell(row_idx, 2).value)
        date_text = normalize_date_text(ws.cell(row_idx, 5).value)
        if not name and not date_text:
            continue
        if not name or not date_text:
            logs.append(LogEntry("WARN", sheet_name, row_idx, "이름 또는 상담일이 없어 건너뜀"))
            counters["skip_missing_common"] += 1
            continue

        birth = normalize_birth_text(ws.cell(row_idx, 3).value)
        gender = normalize_gender(ws.cell(row_idx, 4).value)
        counsel_type = normalize_counsel_type(ws.cell(row_idx, 6).value)
        worker = normalize_text(ws.cell(row_idx, 8).value)
        상담방법 = normalize_text(ws.cell(row_idx, 7).value)

        exact_key, _ = build_keys(date_text, worker, name, birth)
        record = store.create(exact_key, f"{sheet_name}!{row_idx}")
        apply_common_fields(
            record,
            date_text=date_text,
            worker=worker,
            name=name,
            birth=birth,
            gender=gender,
            counsel_type=counsel_type,
        )

        if "방문상담" in 상담방법 or "거리상담" in 상담방법:
            mark_checkbox(record, "서비스_방문상담")
        elif "낮병동" in 상담방법 or "낮병원" in 상담방법:
            append_text(record, "서비스_기타", 상담방법)

        counters["counsel_rows"] += 1


def ingest_hospital_sheet(workbook, store: RecordStore, logs: list[LogEntry], counters: Counter) -> None:
    ws = workbook["입원현황"]
    for row_idx in range(2, ws.max_row + 1):
        name = normalize_text(ws.cell(row_idx, 5).value)
        date_text = normalize_date_text(ws.cell(row_idx, 3).value)
        if not name and not date_text:
            continue
        if not name or not date_text:
            logs.append(LogEntry("WARN", ws.title, row_idx, "이름 또는 입원날짜가 없어 건너뜀"))
            counters["skip_missing_common"] += 1
            continue

        birth = normalize_birth_text(ws.cell(row_idx, 6).value)
        gender = normalize_gender(ws.cell(row_idx, 4).value)
        worker = normalize_text(ws.cell(row_idx, 10).value)
        counsel_type = normalize_counsel_type(ws.cell(row_idx, 8).value)
        hospital_type_raw = normalize_text(ws.cell(row_idx, 7).value)
        hospital_name = normalize_text(ws.cell(row_idx, 9).value)
        hospital_status = normalize_hospital_status(ws.cell(row_idx, 13).value)
        aftercare = normalize_text(ws.cell(row_idx, 16).value)

        exact_key, _ = build_keys(date_text, worker, name, birth)
        record = get_or_create_record(
            store,
            section_name="hospital",
            source_ref=f"{ws.title}!{row_idx}",
            exact_key=exact_key,
        )
        apply_common_fields(
            record,
            date_text=date_text,
            worker=worker,
            name=name,
            birth=birth,
            gender=gender,
            counsel_type=counsel_type,
        )
        set_if_blank(record, "병원_입원유형", normalize_hospital_type(hospital_type_raw))
        set_if_blank(record, "병원_병원명", hospital_name)
        set_if_blank(record, "병원_입원여부", hospital_status)
        set_if_blank(record, "병원_퇴원후연계", aftercare)

        if "보호" in hospital_type_raw:
            logs.append(LogEntry("INFO", ws.title, row_idx, f"입원유형 '{hospital_type_raw}'를 '{record.data['병원_입원유형']}'으로 정규화"))

        counters["hospital_rows"] += 1


def ingest_facility_sheet(workbook, store: RecordStore, logs: list[LogEntry], counters: Counter) -> None:
    ws = workbook["시설입소현황"]
    for row_idx in range(2, ws.max_row + 1):
        name = normalize_text(ws.cell(row_idx, 4).value)
        date_text = normalize_date_text(ws.cell(row_idx, 3).value)
        if not name and not date_text:
            continue
        if not name or not date_text:
            logs.append(LogEntry("WARN", ws.title, row_idx, "이름 또는 날짜가 없어 건너뜀"))
            counters["skip_missing_common"] += 1
            continue

        birth = normalize_birth_text(ws.cell(row_idx, 5).value)
        facility_name = normalize_text(ws.cell(row_idx, 6).value)
        worker = normalize_text(ws.cell(row_idx, 8).value)
        gender = ""
        counsel_type = normalize_counsel_type(ws.cell(row_idx, 7).value)
        facility_type_raw = normalize_text(ws.cell(row_idx, 10).value)
        facility_status = normalize_facility_status(ws.cell(row_idx, 11).value)
        aftercare = normalize_text(ws.cell(row_idx, 14).value)

        exact_key, _ = build_keys(date_text, worker, name, birth)
        record = get_or_create_record(
            store,
            section_name="facility",
            source_ref=f"{ws.title}!{row_idx}",
            exact_key=exact_key,
        )
        apply_common_fields(
            record,
            date_text=date_text,
            worker=worker,
            name=name,
            birth=birth,
            gender=gender,
            counsel_type=counsel_type,
        )
        set_if_blank(record, "시설_입소유형", normalize_facility_type(facility_type_raw))
        set_if_blank(record, "시설_시설명", facility_name)
        set_if_blank(record, "시설_입퇴소여부", facility_status)
        set_if_blank(record, "시설_퇴소후조치", aftercare)

        counters["facility_rows"] += 1


def ingest_referral_sheet(workbook, store: RecordStore, logs: list[LogEntry], counters: Counter) -> None:
    ws = workbook["연계대장"]
    for row_idx in range(5, ws.max_row + 1):
        name = normalize_text(ws.cell(row_idx, 10).value)
        date_text = normalize_date_text(ws.cell(row_idx, 4).value)
        if not name and not date_text:
            continue
        if not name or not date_text:
            logs.append(LogEntry("WARN", ws.title, row_idx, "이름 또는 날짜가 없어 건너뜀"))
            counters["skip_missing_common"] += 1
            continue

        birth = normalize_birth_text(ws.cell(row_idx, 11).value)
        gender = normalize_gender(ws.cell(row_idx, 12).value)
        counsel_type = normalize_counsel_type(ws.cell(row_idx, 13).value)
        worker = normalize_text(ws.cell(row_idx, 8).value)
        referral_org = normalize_text(ws.cell(row_idx, 6).value)
        request = normalize_text(ws.cell(row_idx, 14).value)
        result = normalize_text(ws.cell(row_idx, 16).value)
        detail = normalize_text(ws.cell(row_idx, 17).value)

        exact_key, _ = build_keys(date_text, worker, name, birth)
        record = get_or_create_record(
            store,
            section_name="service",
            source_ref=f"{ws.title}!{row_idx}",
            exact_key=exact_key,
        )
        apply_common_fields(
            record,
            date_text=date_text,
            worker=worker,
            name=name,
            birth=birth,
            gender=gender,
            counsel_type=counsel_type,
        )

        combined_text = " ".join(filter(None, [request, result, detail]))
        public_items = infer_public_service_items(request, result, detail, referral_org)
        extracted_hospital = extract_hospital_name(result, detail)

        if "주거지원" in combined_text or "임시주거" in combined_text:
            mark_checkbox(record, "서비스_주거지원")
        if "외래진료" in combined_text or "외래 진료" in combined_text or "진료의뢰" in combined_text:
            mark_checkbox(record, "서비스_외래진료")
        if "투약" in combined_text:
            mark_checkbox(record, "서비스_투약관리")
        if "물품" in combined_text:
            mark_checkbox(record, "서비스_물품제공")

        if extracted_hospital:
            set_if_blank(record, "서비스_병원명", extracted_hospital)

        if public_items:
            set_if_blank(record, "공적서비스_지원항목", ", ".join(public_items))
            set_if_blank(record, "공적서비스_신청기관", infer_public_service_org(referral_org, detail))
            set_if_blank(record, "공적서비스_지원시기", date_text)

        unstructured = summarize_unstructured_service(result, request, detail)
        if unstructured and not any(record.data[field] for field in [
            "서비스_주거지원",
            "서비스_물품제공",
            "서비스_방문상담",
            "서비스_외래진료",
            "서비스_투약관리",
        ]):
            append_text(record, "서비스_기타", unstructured)
        elif result in {"정보제공", "입원연계", "시설입소", "이탈", "기타"}:
            append_text(record, "서비스_기타", unstructured)

        if not any(normalize_text(record.data.get(field)) for field in SECTION_FIELDS["service"]):
            append_text(record, "서비스_기타", unstructured or request or result or detail)
            logs.append(LogEntry("INFO", ws.title, row_idx, "구조화 매핑이 어려워 서비스_기타로 기록"))

        counters["referral_rows"] += 1


def sort_records(records: Iterable[ImportRecord]) -> list[ImportRecord]:
    def sort_key(record: ImportRecord):
        return (
            normalize_text(record.data.get("공통_날짜")),
            normalize_text(record.data.get("공통_담당자")),
            normalize_text(record.data.get("공통_이름")),
            normalize_text(record.data.get("공통_번호")),
        )

    return sorted(records, key=sort_key)


def autosize_columns(ws) -> None:
    for index, width in enumerate(RAW_COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def style_result_sheet(ws) -> None:
    thin = Side(style="thin", color="D9D9D9")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for title, start_col, end_col, color in GROUP_RANGES:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        cell = ws.cell(1, start_col)
        cell.value = title
        cell.font = header_font
        cell.alignment = center
        cell.fill = PatternFill("solid", fgColor=color)

    for col_idx, header in enumerate(RAW_DETAIL_HEADERS, start=1):
        cell = ws.cell(2, col_idx)
        cell.value = header
        cell.font = header_font
        cell.alignment = center
        group_fill = None
        for _, start_col, end_col, color in GROUP_RANGES:
            if start_col <= col_idx <= end_col:
                group_fill = color
                break
        if group_fill:
            cell.fill = PatternFill("solid", fgColor=group_fill)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(RAW_COLUMN_KEYS)):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row >= 3 and cell.column in (1, 5, 17, 18, 19, 20, 21, 24, 25):
                cell.alignment = center
            elif cell.row >= 3:
                cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(RAW_COLUMN_KEYS))}{ws.max_row}"
    autosize_columns(ws)


def write_result_workbook(
    output_path: Path,
    records: list[ImportRecord],
    logs: list[LogEntry],
    counters: Counter,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "원본데이터_변환결과"

    for row_idx, record in enumerate(records, start=3):
        for col_idx, key in enumerate(RAW_COLUMN_KEYS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=record.data.get(key, ""))

    style_result_sheet(ws)

    summary_ws = wb.create_sheet("변환요약")
    summary_ws.append(["항목", "값"])
    for label, value in [
        ("생성 행 수", len(records)),
        ("상담현황 반영 행", counters["counsel_rows"]),
        ("입원현황 반영 행", counters["hospital_rows"]),
        ("시설입소 반영 행", counters["facility_rows"]),
        ("연계대장 반영 행", counters["referral_rows"]),
        ("누락으로 건너뜀", counters["skip_missing_common"]),
        ("로그 건수", len(logs)),
    ]:
        summary_ws.append([label, value])
    summary_ws.freeze_panes = "A2"

    log_ws = wb.create_sheet("변환로그")
    log_ws.append(["레벨", "원본시트", "원본행", "메시지"])
    for entry in logs:
        log_ws.append([entry.level, entry.source_sheet, entry.source_row, entry.message])
    log_ws.freeze_panes = "A2"
    log_ws.auto_filter.ref = f"A1:D{log_ws.max_row}"
    log_ws.column_dimensions["A"].width = 10
    log_ws.column_dimensions["B"].width = 20
    log_ws.column_dimensions["C"].width = 10
    log_ws.column_dimensions["D"].width = 90

    for sheet in (summary_ws, log_ws):
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DDEBF7")

    wb.save(output_path)


def write_paste_ready_workbook(output_path: Path, records: list[ImportRecord]) -> None:
    wb = Workbook()

    full_ws = wb.active
    full_ws.title = "원본데이터"
    for row_idx, record in enumerate(records, start=3):
        for col_idx, key in enumerate(RAW_COLUMN_KEYS, start=1):
            full_ws.cell(row=row_idx, column=col_idx, value=record.data.get(key, ""))
    style_result_sheet(full_ws)

    paste_ws = wb.create_sheet("A3붙여넣기용")
    for row_idx, record in enumerate(records, start=1):
        for col_idx, key in enumerate(RAW_COLUMN_KEYS, start=1):
            paste_ws.cell(row=row_idx, column=col_idx, value=record.data.get(key, ""))
    autosize_columns(paste_ws)

    guide_ws = wb.create_sheet("사용안내")
    guide_ws.append(["순서", "설명"])
    guide_ws.append([1, "원본데이터 시트는 헤더까지 포함된 최종본입니다. 새 검토용으로 사용하세요."])
    guide_ws.append([2, "A3붙여넣기용 시트는 헤더가 없습니다. 전체 범위를 복사해 구글시트 원본데이터 A3 셀에 붙여넣으세요."])
    guide_ws.append([3, "붙여넣기 후 데이터 자동화 > 전체 재구성 (동기화)를 실행하세요."])
    guide_ws.column_dimensions["A"].width = 10
    guide_ws.column_dimensions["B"].width = 90
    for cell in guide_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDEBF7")

    wb.save(output_path)


def write_paste_ready_tsv(output_path: Path, records: list[ImportRecord]) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as fp:
        writer = csv.writer(fp, delimiter="\t")
        for record in records:
            writer.writerow([record.data.get(key, "") for key in RAW_COLUMN_KEYS])


def detect_input_file(explicit_path: str | None) -> Path:
    if explicit_path:
        path = Path(explicit_path).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"입력 파일을 찾을 수 없습니다: {path}")
        return path

    candidates = sorted(BASE_DIR.glob("2026*정신건강팀 실적.xlsx"))
    if not candidates:
        raise FileNotFoundError("python-tools 폴더에서 '2026 지역별 일일실적 정신건강팀 실적.xlsx' 파일을 찾지 못했습니다.")
    return candidates[0]


def detect_output_file(explicit_path: str | None, input_path: Path) -> Path:
    if explicit_path:
        return Path(explicit_path).expanduser().resolve()
    return input_path.with_name(f"{input_path.stem}_원본데이터변환.xlsx")


def build_derivative_output_paths(main_output_path: Path) -> tuple[Path, Path]:
    stem = main_output_path.stem.replace("_원본데이터변환", "")
    parent = main_output_path.parent
    paste_workbook = parent / f"{stem}_구글시트붙여넣기용.xlsx"
    paste_tsv = parent / f"{stem}_구글시트붙여넣기용.tsv"
    return paste_workbook, paste_tsv


def main() -> None:
    parser = argparse.ArgumentParser(description="과거 정신건강팀 엑셀을 현재 원본데이터 형식으로 변환합니다.")
    parser.add_argument("--input", help="원본 지역별 일일실적 엑셀 경로")
    parser.add_argument("--output", help="변환 결과 엑셀 경로")
    args = parser.parse_args()

    input_path = detect_input_file(args.input)
    output_path = detect_output_file(args.output, input_path)

    workbook = load_source_workbook(input_path)
    store = RecordStore()
    logs: list[LogEntry] = []
    counters: Counter = Counter()

    ingest_counsel_sheet(workbook, "상담현황(2025)", store, logs, counters)
    if "상담현황(2026)" in workbook.sheetnames:
        ingest_counsel_sheet(workbook, "상담현황(2026)", store, logs, counters)
    ingest_hospital_sheet(workbook, store, logs, counters)
    ingest_facility_sheet(workbook, store, logs, counters)
    ingest_referral_sheet(workbook, store, logs, counters)

    sorted_records = sort_records(store.records)
    for index, record in enumerate(sorted_records, start=1):
        record.data["공통_번호"] = f"LEGACY-{index:05d}"
    write_result_workbook(output_path, sorted_records, logs, counters)
    paste_workbook_path, paste_tsv_path = build_derivative_output_paths(output_path)
    write_paste_ready_workbook(paste_workbook_path, sorted_records)
    write_paste_ready_tsv(paste_tsv_path, sorted_records)

    print(f"입력 파일: {input_path}")
    print(f"출력 파일: {output_path}")
    print(f"붙여넣기용 엑셀: {paste_workbook_path}")
    print(f"붙여넣기용 TSV: {paste_tsv_path}")
    print(f"생성 행 수: {len(sorted_records)}")
    print(f"상담현황 반영: {counters['counsel_rows']}")
    print(f"입원현황 반영: {counters['hospital_rows']}")
    print(f"시설입소 반영: {counters['facility_rows']}")
    print(f"연계대장 반영: {counters['referral_rows']}")
    print(f"건너뜀: {counters['skip_missing_common']}")
    print(f"로그 건수: {len(logs)}")


if __name__ == "__main__":
    main()
