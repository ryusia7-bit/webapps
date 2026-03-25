import pandas as pd
import numpy as np
import datetime
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(BASE_DIR, "정신건강팀 실적2026.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "정신건강팀_실적결과대시보드.xlsx")

def clean_val(v):
    if pd.isna(v):
        return ""
    if isinstance(v, datetime.datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()

def is_truthy(v):
    if pd.isna(v):
        return False
    v_str = str(v).strip().lower()
    return v_str in ["true", "v", "o", "1", "yes", "y", "체크", "선택"]

def main():
    print(f"엑셀 데이터 읽는 중: {INPUT_FILE}")
    
    # 헤더 병합 문제가 있을 수 있으므로 None으로 읽고 직접 헤더를 조립합니다.
    df_raw = pd.read_excel(INPUT_FILE, sheet_name="기본DB", header=None)
    
    # 0행: 그룹 헤더 (공통, 병원, 시설 등) - ffill 적용
    r0 = df_raw.iloc[0].ffill()
    # 1행: 실제 헤더 (번호, 날짜, 주거지원 등)
    r1 = df_raw.iloc[1]
    
    headers = []
    for h1, h2 in zip(r0, r1):
        v1 = str(h1).strip() if not pd.isna(h1) else ""
        v2 = str(h2).strip() if not pd.isna(h2) else ""
        
        if v1 and v1 != "nan" and not v1.startswith("Unnamed"):
            headers.append(f"{v1}_{v2}" if v2 and v2 != "nan" else v1)
        else:
            headers.append(v2 if v2 and v2 != "nan" else "UNKNOWN")

    # 데이터 프레임 슬라이싱
    df = df_raw.iloc[2:].copy()
    df.columns = headers
    
    # 고유번호 기준 (번호 열)
    # 엑셀 형식이 'Unnamed'라면 그냥 마지막 항목을 보고 때려 잡습니다
    # 혹은 안전하게 '번호' 열 구하기
    no_col = [c for c in headers if c.endswith("번호") and c.startswith("공통")]
    if no_col:
        no_col = no_col[0]
    else:
        # Fallback
        for c in headers:
            if "번호" in c:
                no_col = c
                break
                
    # 결과 배열
    base_rows = []
    hospital_rows = []
    facility_rows = []
    service_rows = []
    error_rows = []

    # 순회하면서 Apps Script와 동일한 로직 적용
    for idx, row in df.iterrows():
        row_num = idx + 1 # 엑셀 행 번호 (1-based)
        
        date_raw = row.get("공통_날짜", "")
        name = row.get("공통_이름", "")
        
        date_str = clean_val(date_raw)
        name_str = clean_val(name)
        
        if not date_str and not name_str:
            continue # 완전히 빈 줄 패스
            
        record_id = str(row.get(no_col, row_num)).strip()
        if not record_id or record_id == "nan" or record_id == "None":
            record_id = str(row_num)
            
        worker = clean_val(row.get("공통_담당자", ""))
        birth_raw = clean_val(row.get("공통_생년월일", ""))
        gender = clean_val(row.get("공통_성별", ""))
        counsel_type = clean_val(row.get("공통_상담유형", ""))
        
        # 기본정보
        base_rows.append({
            "고유번호": record_id,
            "원본행": row_num,
            "날짜": date_str,
            "담당자": worker,
            "이름": name_str,
            "생년월일_원본": birth_raw,
            "성별": gender,
            "상담유형": counsel_type,
            "유효성": bool(date_str and name_str)
        })
        
        # 병원기록
        hosp_type = clean_val(row.get("병원_입원유형", ""))
        hosp_name = clean_val(row.get("병원_병원명", ""))
        hosp_status = clean_val(row.get("병원_입원여부", ""))
        hosp_action = clean_val(row.get("병원_퇴원후연계", ""))
        
        if hosp_type or hosp_name or hosp_status or hosp_action:
            hospital_rows.append({
                "고유번호": record_id,
                "원본행": row_num,
                "입원유형": hosp_type,
                "병원명": hosp_name,
                "입원여부": hosp_status,
                "퇴원후연계": hosp_action
            })
            
        # 시설기록
        fac_type = clean_val(row.get("시설_입소유형", ""))
        fac_name = clean_val(row.get("시설_시설명", ""))
        fac_status = clean_val(row.get("시설_입퇴소여부", ""))
        fac_action = clean_val(row.get("시설_퇴소후조치", ""))
        
        if fac_type or fac_name or fac_status or fac_action:
            facility_rows.append({
                "고유번호": record_id,
                "원본행": row_num,
                "입소유형": fac_type,
                "시설명": fac_name,
                "입퇴소여부": fac_status,
                "퇴소후조치": fac_action
            })
            
        # 서비스기록
        svc_types = [
            ("서비스_주거지원", "주거지원"),
            ("서비스_물품제공", "물품제공"),
            ("서비스_방문상담", "방문상담"),
            ("서비스_외래진료", "외래진료"),
            ("서비스_투약관리", "투약관리"),
            ("서비스_공적서비스", "공적서비스")
        ]
        
        svc_org = clean_val(row.get("서비스_병원명", ""))
        seq = 1
        
        for key, s_type in svc_types:
            val = row.get(key, "")
            if is_truthy(val):
                detail = ""
                s_org = svc_org
                
                if s_type == "공적서비스":
                    support_item = clean_val(row.get("공적서비스_지원항목", ""))
                    detail = support_item
                    applicant_org = clean_val(row.get("공적서비스_신청기관", ""))
                    if applicant_org:
                        s_org = applicant_org
                    support_date = clean_val(row.get("공적서비스_지원시기", ""))
                    if support_date:
                        detail += f" / 지원시기: {support_date}" if detail else f"지원시기: {support_date}"
                
                service_rows.append({
                    "서비스번호": f"{record_id}_{seq}",
                    "고유번호": record_id,
                    "원본행": row_num,
                    "서비스유형": s_type,
                    "서비스상세": detail,
                    "관련기관명": s_org,
                    "서비스일자": date_str,
                    "담당자": worker,
                    "이름": name_str
                })
                seq += 1
                
        # 서비스_기타
        svc_etc = clean_val(row.get("서비스_기타", ""))
        if svc_etc:
             service_rows.append({
                    "서비스번호": f"{record_id}_{seq}",
                    "고유번호": record_id,
                    "원본행": row_num,
                    "서비스유형": "기타",
                    "서비스상세": svc_etc,
                    "관련기관명": svc_org,
                    "서비스일자": date_str,
                    "담당자": worker,
                    "이름": name_str
                })
             seq += 1
                
    # 엑셀 형태로 저장 준비
    print("엑셀 시트별로 저장중...")
    
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        df_base = pd.DataFrame(base_rows)
        df_base.to_excel(writer, sheet_name="기본정보", index=False)
        
        df_hospital = pd.DataFrame(hospital_rows)
        df_hospital.to_excel(writer, sheet_name="병원기록", index=False)
        
        df_fac = pd.DataFrame(facility_rows)
        df_fac.to_excel(writer, sheet_name="시설기록", index=False)
        
        df_svc = pd.DataFrame(service_rows)
        df_svc.to_excel(writer, sheet_name="서비스기록", index=False)

        workbook = writer.book
        
        from openpyxl.styles import PatternFill, Font, Alignment
        header_fill = PatternFill(start_color="DDEBF7", fill_type="solid")
        header_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        # 한글 길이를 고려해 약간 가중치 보정
                        val_str = str(cell.value)
                        v_len = len(val_str) + sum(1 for c in val_str if ord(c) > 127) * 0.5
                        if v_len > max_length:
                            max_length = v_len # 문자열 길이 측정
                    except:
                        pass
                
                adjusted_width = (max_length + 2) * 1.1
                ws.column_dimensions[column].width = min(adjusted_width, 60) # 최대넓이 60 제한

    print(f"변환 완료! 결과 파일: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
